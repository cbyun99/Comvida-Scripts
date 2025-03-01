import openpyxl as pyxl
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.worksheet.pagebreak import Break
from openpyxl.worksheet.cell_range import CellRange
from datetime import datetime, timedelta

def attendance_log(file_path="Reports/SSAttendanceLog.xlsx", save_file_path=None, sheet_name="SSAttendanceLog", offset=1):
    # Load the workbook and select the worksheet
    wb = pyxl.load_workbook(file_path)
    ws = wb[sheet_name]
    
    #change worksheet margins to minimum
    ws.page_margins.top = 0.1
    ws.page_margins.bottom = 0.1    

    # Define the array of values to check in column D
    values_array = [
        "ACT-D", "DWK1", "DWK2", "DWK3", "HK-D1", "HK-D2", "HK-D3", "HK-E",
        "FSC1", "FSC2", "FSD1", "FSD2", "FSD4", "FSE1", "FSE2", "HOME-WKD",
        "HOME - E", "HOME-D", "HRSCH", "Ls1", "Ls2", "LAD1", "LAD2"
    ]

    location = "TH" if str(ws.cell(row=3, column=4).value).startswith("1") else "VH" if str(ws.cell(row=3, column=4).value).startswith("C") else "TC"

    #Combine Rows that are doubled due to OT and shifts with BANKTK
    max_row = ws.max_row
    deleted_rows = 0
    i = 2
    while i < max_row - 1:
        b_i = ws[f"B{i}"].value #listname
        b_next = ws[f"B{i+1}"].value 
        d_i = ws[f"D{i}"].value #shiftcode
        d_next = ws[f"D{i+1}"].value 
        n_i = ws[f"N{i}"].value #starttime
        n_next = ws[f"N{i+1}"].value 
        o_i= ws[f"O{i}"].value  #endtime
        o_next = ws[f"O{i+1}"].value 
        
        if b_i == b_next and d_i == d_next: #if duplicate shift (due to OT)
            p_i = ws[f"P{i}"].value if ws[f"P{i}"].value is not None else 0
            p_next = ws[f"P{i+1}"].value if ws[f"P{i+1}"].value is not None else 0
            p_val = p_i + p_next
            ws[f"P{i}"].value = p_val
            ws[f"N{i}"].value = min(n_i, n_next)
            ws[f"O{i}"].value = max(o_i, o_next)
            print(d_i + " " + b_i)
            ws.delete_rows(i+1)
            deleted_rows += 1
            i += 1
        elif d_i == "BANKTK": #if shift is BANKTK
            ws.delete_rows(i)
            print(b_i + " row " + i + " deleted (BANKTK)" )
            deleted_rows += 1
            i+=1
        else:
            i += 1

    # Move SS to the bottom for TH location
    if location =="TH":
        # Find last row in column A
        last_row = ws.max_row
        # Move matched rows to the bottom
        move_rows = []
        for i in range(2, last_row + 1):  # Assuming row 1 has headers
            cell_value = ws.cell(row=i, column=4).value  # Column D (4th column)
            if cell_value in values_array:
                move_rows.append([(ws.cell(row=i, column=j).value, ws.cell(row=i, column=j).number_format) for j in range(1, ws.max_column + 1)])

        # Remove matched rows
        for row in reversed(range(2, last_row + 1)):
            if ws.cell(row=row, column=4).value in values_array:
                ws.delete_rows(row)

        # Append the moved rows at the bottom with their original formatting
        for row in move_rows:
            new_row = ws.max_row + 1 
            for col, (value, num_format) in enumerate(row, start=1):
                cell = ws.cell(row=new_row, column=col, value=value)
                cell.number_format = num_format

        # Find "RNN" in column D and add a page break
        row_height = .21*72
        for i in range(2, ws.max_row + 1 ):
            if ws.cell(row=i, column=4).value == "RNN":
                first_page_max = i + 1 
                if first_page_max > 52: # If first page is too long, decrease row height
                    row_height -= ((first_page_max - 49) // 3)*.72
                    for row in range(2, first_page_max):
                        ws.row_dimensions[row].height = row_height
                ws.row_breaks.append(Break(id=i+1))
                if deleted_rows > 1:
                    ws.delete_rows(i+1)
                break

    # Hide specific columns
    for col in ["A", "E", "F", "G", "H", "I", "J", "K", "L", "M", "Q"]:
        ws.column_dimensions[col].hidden = True

    # Adjust column widths
    column_widths = {"B": 34.29,"C": 13.57, "D": 9, "N": 9, "O": 8.43, "P": 6.14}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Insert title row
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=2)
    title_cell.value = f"Attendance Log {location} {(datetime.today() + timedelta (days=offset)).strftime('%B %d, %Y')}"
    title_cell.font = Font(size=14, bold=True)

    # Merge and center the title row (B1 to P1)
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=16)
    title_cell.alignment = pyxl.styles.Alignment(horizontal="center", vertical="center")

    min_col = 2
    max_col = 16

    # Altered borders for TH location
    if location =="TH":
        border_ranges = [CellRange("B1:P" + str(first_page_max)), 
                             CellRange("B" + str(first_page_max + 1) + ":P" + str(ws.max_row))]
    else:
        border_ranges = [CellRange("B1:P" + str(ws.max_row))]
    border_ranges.extend([CellRange("B1:P1"), CellRange("B2:P2")])

    for border_range in border_ranges:
        for row, col in border_range.cells:
            top = Side(style="medium") if (row, col) in border_range.top else None
            left = Side(style="medium") if (row, col) in border_range.left else None
            right = Side(style="medium") if (row, col) in border_range.right else None
            bottom = Side(style="medium") if (row, col) in border_range.bottom else None
            ws.cell(row, col).border = Border(left, right, top, bottom, outline=True)

    # Underline Unit Borders
    for row in range(1, ws.max_row -1):
        d1 = ws[f"D{row}"].value
        d2 = ws[f"D{row+1}"].value
        if d1 and d2 and isinstance(d1, str) and isinstance(d2, str):
            if d1[:1] != d2[:1] and d1[:1].isdigit():
                ws.cell(row, 2).border = Border(bottom=Side(style="thin"), 
                                                    left=Side(style="medium"))
                ws.cell(row, 16).border = Border(bottom=Side(style="thin"), 
                                                    right=Side(style="medium"))
                for col in range(min_col + 1, max_col):
                    ws.cell(row, col).border = Border(bottom=Side(style="thin"))

    # Conditional Formatting - highlight RNs and LPNs
    gray_fill = PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid")
    for i in range(2, ws.max_row + 1):
        cell_value = str(ws.cell(row=i, column=4).value)
        if any(keyword in cell_value for keyword in ["LPN", "RCC", "RN"]):
            for col in range(2, 17):  # Apply to columns B to P
                ws.cell(row=i, column=col).fill = gray_fill

    # Save workbook
    if save_file_path == None:
        save_file_path = "processed_" + file_path
    wb.save(save_file_path)

def main(): 
    attendance_log("example alog 3.xlsx")


if __name__ == "__main__":
    main()